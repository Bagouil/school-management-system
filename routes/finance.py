from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file,jsonify,request
from models.finance import FeeCategory, StudentFee, Expense
from models.student import Student
from models.class_ import Class
from database.db_config import get_db
from datetime import datetime, timedelta
from utils.pdf_receipt import PDFReceiptGenerator
from utils.excel_reports import ExcelReportGenerator
from utils.pdf_reports import PDFReportGenerator
from models.installment import InstallmentPlan, StudentInstallment
from utils.permission_decorator import role_required

import pandas as pd
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import csv
finance_bp = Blueprint('finance', __name__)

# ==================== DASHBOARD ====================
@finance_bp.route('/dashboard')
@role_required('admin', 'accountant')
def dashboard():
    """Finance dashboard"""
    
    db = get_db()
    cursor = db.cursor()
    
    # Get current period
    period = request.args.get('period', 'month')
    today = datetime.now()
    
    # Set date ranges
    if period == 'today':
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif period == 'week':
        start_date = (today - timedelta(days=7)).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    elif period == 'month':
        start_date = today.strftime('%Y-%m-01')
        end_date = today.strftime('%Y-%m-%d')
    else:  # year
        start_date = today.strftime('%Y-01-01')
        end_date = today.strftime('%Y-%m-%d')
    
    # Get totals - convert to float immediately
    cursor.execute("SELECT ISNULL(SUM(amount_paid), 0) FROM FeePayments WHERE payment_date BETWEEN ? AND ?", 
                  (start_date, end_date))
    total_collected = float(cursor.fetchone()[0] or 0)
    
    cursor.execute("SELECT ISNULL(SUM(amount), 0) FROM Expenses WHERE expense_date BETWEEN ? AND ?", 
                  (start_date, end_date))
    total_expenses = float(cursor.fetchone()[0] or 0)
    
    net_income = total_collected - total_expenses
    
    # Pending fees
    cursor.execute("""
    SELECT ISNULL(SUM(sf.amount - ISNULL(sf.discount_amount, 0) - 
                   ISNULL(fp.total_paid, 0)), 0)
    FROM StudentFees sf
    LEFT JOIN (
        SELECT student_fee_id, SUM(amount_paid) as total_paid
        FROM FeePayments
        GROUP BY student_fee_id
    ) fp ON sf.student_fee_id = fp.student_fee_id
    WHERE sf.status IN ('pending', 'partial')
    """)
    pending_fees = float(cursor.fetchone()[0] or 0)
    
    # Count students with pending fees
    cursor.execute("""
    SELECT COUNT(DISTINCT sf.student_id)
    FROM StudentFees sf
    LEFT JOIN (
        SELECT student_fee_id, SUM(amount_paid) as total_paid
        FROM FeePayments
        GROUP BY student_fee_id
    ) fp ON sf.student_fee_id = fp.student_fee_id
    WHERE sf.status IN ('pending', 'partial')
    AND (sf.amount - ISNULL(sf.discount_amount, 0) - ISNULL(fp.total_paid, 0)) > 0
    """)
    pending_count = int(cursor.fetchone()[0] or 0)
    
    # Growth percentages - ensure they are floats
    collection_growth = 12.5
    expense_growth = 8.3
    income_percentage = 15.7
    
    # Collection percentage
    cursor.execute("""
    SELECT 
        ISNULL(SUM(sf.amount - ISNULL(sf.discount_amount, 0)), 0) as total_fees,
        ISNULL(SUM(ISNULL(fp.total_paid, 0)), 0) as total_paid
    FROM StudentFees sf
    LEFT JOIN (
        SELECT student_fee_id, SUM(amount_paid) as total_paid
        FROM FeePayments
        GROUP BY student_fee_id
    ) fp ON sf.student_fee_id = fp.student_fee_id
    """)
    fees_data = cursor.fetchone()
    total_fees = float(fees_data[0] or 1)
    total_paid_fees = float(fees_data[1] or 0)
    collection_percentage = round((total_paid_fees / total_fees) * 100, 1) if total_fees > 0 else 0.0
    
    # Get recent payments
    cursor.execute("""
    SELECT TOP 5
        CONCAT(s.first_name_ar, ' ', s.last_name_ar) as student_name,
        fp.payment_date,
        fp.receipt_number,
        fp.amount_paid,
        u.username
    FROM FeePayments fp
    JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
    JOIN Students s ON sf.student_id = s.student_id
    JOIN Users u ON fp.received_by = u.user_id
    ORDER BY fp.payment_date DESC
    """)
    
    recent_payments = []
    for row in cursor.fetchall():
        recent_payments.append({
            'student_name': row[0],
            'date': row[1].strftime('%Y-%m-%d') if row[1] else '',
            'receipt': row[2],
            'amount': float(row[3]),
            'user': row[4]
        })
    
    # Get recent expenses
    cursor.execute("""
    SELECT TOP 5
        description,
        expense_date,
        amount,
        expense_category
    FROM Expenses
    ORDER BY expense_date DESC
    """)
    
    recent_expenses = []
    for row in cursor.fetchall():
        recent_expenses.append({
            'description': row[0],
            'date': row[1].strftime('%Y-%m-%d') if row[1] else '',
            'amount': float(row[2]),
            'category': row[3]
        })
    
    # Get top debtors
    cursor.execute("""
    SELECT TOP 5
        s.student_id,
        CONCAT(s.first_name_ar, ' ', s.last_name_ar) as name,
        c.class_name_ar as class,
        SUM(sf.amount - ISNULL(sf.discount_amount, 0) - ISNULL(fp.total_paid, 0)) as amount
    FROM StudentFees sf
    JOIN Students s ON sf.student_id = s.student_id
    JOIN Classes c ON s.current_class_id = c.class_id
    LEFT JOIN (
        SELECT student_fee_id, SUM(amount_paid) as total_paid
        FROM FeePayments
        GROUP BY student_fee_id
    ) fp ON sf.student_fee_id = fp.student_fee_id
    WHERE sf.status IN ('pending', 'partial')
    GROUP BY s.student_id, s.first_name_ar, s.last_name_ar, c.class_name_ar
    HAVING SUM(sf.amount - ISNULL(sf.discount_amount, 0) - ISNULL(fp.total_paid, 0)) > 0
    ORDER BY amount DESC
    """)
    
    top_debtors = []
    for row in cursor.fetchall():
        top_debtors.append({
            'id': row[0],
            'name': row[1],
            'class': row[2],
            'amount': float(row[3])
        })
    
    # Chart data
    chart_labels = []
    income_data = []
    expense_data = []
    
    for i in range(6, -1, -1):
        date = (today - timedelta(days=i)).strftime('%Y-%m-%d')
        chart_labels.append((today - timedelta(days=i)).strftime('%a'))
        
        cursor.execute("SELECT ISNULL(SUM(amount_paid), 0) FROM FeePayments WHERE payment_date = ?", (date,))
        income_data.append(float(cursor.fetchone()[0]))
        
        cursor.execute("SELECT ISNULL(SUM(amount), 0) FROM Expenses WHERE expense_date = ?", (date,))
        expense_data.append(float(cursor.fetchone()[0]))
    
    cursor.close()
    
    # Prepare context with proper types
    context = {
        # Raw numeric values for calculations
        'total_collected_raw': total_collected,
        'total_expenses_raw': total_expenses,
        'net_income_raw': net_income,
        'pending_fees_raw': pending_fees,
        'pending_count': pending_count,
        'collection_growth': collection_growth,
        'expense_growth': expense_growth,
        'income_percentage': income_percentage,
        'collection_percentage': collection_percentage,
        
        # Formatted values for display
        'total_collected': f"{total_collected:,.0f}",
        'total_expenses': f"{total_expenses:,.0f}",
        'net_income': f"{net_income:,.0f}",
        'pending_fees': f"{pending_fees:,.0f}",
        
        # Other data
        'recent_payments': recent_payments,
        'recent_expenses': recent_expenses,
        'top_debtors': top_debtors,
        'chart_labels': chart_labels,
        'income_data': income_data,
        'expense_data': expense_data,
        'period': period
    }
    
    return render_template('finance/dashboard.html', **context)

# ==================== FEE CATEGORIES ====================
@finance_bp.route('/fee-categories')
@role_required('admin', 'accountant')
def fee_categories():
    """List fee categories"""

    
    db = get_db()
    cursor = db.cursor()
    
    # Simple query without the new columns
    try:
        cursor.execute("""
        SELECT 
            fc.fee_category_id,
            fc.category_name_ar,
            fc.category_name_en,
            fc.description,
            fc.amount,
            fc.is_annual,
            fc.is_mandatory,
            COUNT(DISTINCT sf.student_id) as student_count,
            ISNULL(SUM(fp.amount_paid), 0) as collected
        FROM FeeCategories fc
        LEFT JOIN StudentFees sf ON fc.fee_category_id = sf.fee_category_id
        LEFT JOIN FeePayments fp ON sf.student_fee_id = fp.student_fee_id
        GROUP BY fc.fee_category_id, fc.category_name_ar, fc.category_name_en, 
                 fc.description, fc.amount, fc.is_annual, fc.is_mandatory
        ORDER BY fc.is_mandatory DESC, fc.category_name_ar
        """)
    except Exception as e:
        # Fallback even simpler query if the above fails
        cursor.execute("SELECT * FROM FeeCategories ORDER BY category_name_ar")
    
    columns = [column[0] for column in cursor.description]
    categories = []
    total_categories = 0
    mandatory_count = 0
    annual_count = 0
    
    for row in cursor.fetchall():
        category = dict(zip(columns, row))
        # Add default values for missing fields
        category['category_code'] = ''
        category['is_active'] = 1
        categories.append(category)
        
        total_categories += 1
        if category.get('is_mandatory'):
            mandatory_count += 1
        if category.get('is_annual'):
            annual_count += 1
    
    # Get classes for assign modal
    cursor.execute("SELECT class_id, class_name_ar FROM Classes ORDER BY class_name_ar")
    classes = []
    for row in cursor.fetchall():
        classes.append({'class_id': row[0], 'class_name_ar': row[1]})
    
    cursor.close()
    
    return render_template('finance/fee_categories.html',
                         categories=categories,
                         total_categories=total_categories,
                         mandatory_count=mandatory_count,
                         annual_count=annual_count,
                         active_categories=total_categories,  # All are active for now
                         classes=classes)
@finance_bp.route('/fee-categories/add', methods=['POST'])
@role_required('admin', 'accountant')
def add_fee_category():  # This function name must match what's used in the template
    """Add new fee category"""
    
    try:
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("""
        INSERT INTO FeeCategories (
            category_name_ar, category_name_en, description, amount, 
            is_annual, is_mandatory
        ) OUTPUT INSERTED.fee_category_id
        VALUES (?, ?, ?, ?, ?, ?)
        """, (
            request.form['category_name_ar'],
            request.form['category_name_en'],
            request.form.get('description', ''),
            float(request.form['amount']),
            1 if 'is_annual' in request.form else 0,
            1 if 'is_mandatory' in request.form else 0
        ))
        
        category_id = cursor.fetchone()[0]
        db.commit()
        cursor.close()
        
        flash('Fee category added successfully', 'success')
        
    except Exception as e:
        flash(f'Error adding fee category: {str(e)}', 'danger')
        db.rollback()
    
    return redirect(url_for('finance.fee_categories'))
@finance_bp.route('/fee-categories/<int:category_id>/data')
@role_required('admin', 'accountant')
def get_category_data(category_id):
    """Get category data for editing"""

    
    category = FeeCategory.get_by_id(category_id)
    if category:
        return jsonify(category)
    return jsonify({'error': 'Category not found'}), 404

@finance_bp.route('/fee-categories/<int:category_id>/edit', methods=['POST'])
@role_required('admin', 'accountant')
def edit_fee_category(category_id):
    """Edit fee category"""

    
    try:
        data = {
            'category_name_ar': request.form['category_name_ar'],
            'category_name_en': request.form['category_name_en'],
            'description': request.form.get('description', ''),
            'amount': float(request.form['amount']),
            'category_code': request.form.get('category_code', ''),
            'is_annual': 'is_annual' in request.form,
            'is_mandatory': 'is_mandatory' in request.form,
            'is_active': 'is_active' in request.form
        }
        
        if FeeCategory.update(category_id, data):
            flash('Fee category updated successfully', 'success')
        else:
            flash('No changes were made', 'info')
            
    except Exception as e:
        flash(f'Error updating fee category: {str(e)}', 'danger')
    
    return redirect(url_for('finance.fee_categories'))

@finance_bp.route('/fee-categories/<int:category_id>/delete', methods=['POST'])
@role_required('admin', 'accountant')
def delete_fee_category(category_id):
    """Delete fee category"""
    
    try:
        if FeeCategory.delete(category_id):
            return jsonify({'success': True})
        else:
            return jsonify({'error': 'Cannot delete category in use'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@finance_bp.route('/fee-categories/assign', methods=['POST'])
@role_required('admin', 'accountant')
def assign_category_to_students():
    """Assign fee category to selected students"""

    
    try:
        data = request.get_json()
        category_id = data['category_id']
        student_ids = data['students']
        
        db = get_db()
        cursor = db.cursor()
        
        # Get current academic year
        cursor.execute("SELECT year_id FROM AcademicYears WHERE is_current = 1")
        year_id = cursor.fetchone()[0]
        
        # Get category amount
        cursor.execute("SELECT amount FROM FeeCategories WHERE fee_category_id = ?", (category_id,))
        amount = cursor.fetchone()[0]
        
        # Assign to each student
        for student_id in student_ids:
            cursor.execute("""
            INSERT INTO StudentFees (student_id, fee_category_id, academic_year_id, amount, due_date, status)
            VALUES (?, ?, ?, ?, DATEADD(month, 1, GETDATE()), 'pending')
            """, (student_id, category_id, year_id, amount))
        
        db.commit()
        cursor.close()
        
        return jsonify({'success': True, 'count': len(student_ids)})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@finance_bp.route('/api/students/active')
def get_active_students():
    """Get all active students for API"""
    if 'user_id' not in session:
        return jsonify({'error': 'Access denied'}), 403
    
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("""
    SELECT 
        s.student_id,
        s.student_number,
        s.first_name_ar,
        s.last_name_ar,
        c.class_id,
        c.class_name_ar
    FROM Students s
    JOIN Classes c ON s.current_class_id = c.class_id
    WHERE s.status = 'active'
    ORDER BY s.first_name_ar
    """)
    
    columns = [column[0] for column in cursor.description]
    students = []
    for row in cursor.fetchall():
        students.append(dict(zip(columns, row)))
    
    cursor.close()
    return jsonify(students)

# ==================== STUDENT FEES ====================
@role_required('admin', 'accountant')
@finance_bp.route('/student-fees')
def student_fees():
    """List all student fees"""

    
    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Number of fees per page
    
    db = get_db()
    cursor = db.cursor()
    
    # Get total count for pagination
    cursor.execute("SELECT COUNT(*) FROM StudentFees")
    total = cursor.fetchone()[0]
    
    # Get fees with pagination
    cursor.execute("""
    SELECT 
        sf.student_fee_id,
        s.student_id,
        s.student_number,
        s.first_name_ar as student_name_ar,
        s.first_name_en as student_name_en,
        CONCAT(s.first_name_ar, ' ', s.last_name_ar) as student_name,
        c.class_id,
        c.class_name_ar as class_name,
        fc.category_name_ar as fee_type,
        sf.amount,
        sf.discount_amount as discount,
        ISNULL(SUM(fp.amount_paid), 0) as paid,
        sf.due_date,
        sf.status,
        CASE WHEN sf.due_date < GETDATE() AND sf.status != 'paid' THEN 1 ELSE 0 END as is_overdue
    FROM StudentFees sf
    JOIN Students s ON sf.student_id = s.student_id
    JOIN Classes c ON s.current_class_id = c.class_id
    JOIN FeeCategories fc ON sf.fee_category_id = fc.fee_category_id
    LEFT JOIN FeePayments fp ON sf.student_fee_id = fp.student_fee_id
    GROUP BY sf.student_fee_id, s.student_id, s.student_number, s.first_name_ar, s.last_name_ar,
             s.first_name_en, c.class_id, c.class_name_ar, fc.category_name_ar,
             sf.amount, sf.discount_amount, sf.due_date, sf.status
    ORDER BY sf.due_date DESC
    OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
    """, ((page - 1) * per_page, per_page))
    
    columns = [column[0] for column in cursor.description]
    fees = []
    
    for row in cursor.fetchall():
        fee = dict(zip(columns, row))
        fee['balance'] = fee['amount'] - fee['discount'] - fee['paid']
        fees.append(fee)
    
    # Get classes for filter
    cursor.execute("SELECT class_id, class_name_ar FROM Classes ORDER BY class_name_ar")
    classes = []
    for row in cursor.fetchall():
        classes.append({'class_id': row[0], 'class_name_ar': row[1]})
    
    cursor.close()
    
    # Pagination data
    total_pages = (total + per_page - 1) // per_page
    current_page = page
    
    return render_template('finance/student_fees.html',
                         fees=fees,
                         classes=classes,
                         total_pages=total_pages,
                         current_page=current_page)

@finance_bp.route('/student/<int:student_id>/fees')
def student_fees_detail(student_id):
    """View student fees"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    from models.student import Student
    from models.finance import FeeCategory
    from models.installment import StudentInstallment
    from datetime import date
    
    student = Student.get_by_id(student_id)
    if not student:
        flash('Student not found', 'danger')
        return redirect(url_for('students.list_students'))
    
    db = get_db()
    cursor = db.cursor()
    
    # Get current academic year
    cursor.execute("SELECT year_id FROM AcademicYears WHERE is_current = 1")
    current_year = cursor.fetchone()
    year_id = current_year[0] if current_year else None
    
    # Get fees with payment details - REMOVED payment_count from query
    cursor.execute("""
    SELECT 
        sf.*,
        fc.category_name_ar,
        fc.category_name_en,
        ISNULL(SUM(fp.amount_paid), 0) as total_paid,
        CASE WHEN sf.due_date < GETDATE() AND sf.status != 'paid' THEN 1 ELSE 0 END as is_overdue
    FROM StudentFees sf
    JOIN FeeCategories fc ON sf.fee_category_id = fc.fee_category_id
    LEFT JOIN FeePayments fp ON sf.student_fee_id = fp.student_fee_id
    WHERE sf.student_id = ?
    GROUP BY sf.student_fee_id, sf.student_id, sf.fee_category_id, sf.academic_year_id,
             sf.amount, sf.discount_amount, sf.due_date, sf.status, sf.notes,
             sf.installment_plan_id, sf.allow_partial_payment,
             fc.category_name_ar, fc.category_name_en
    ORDER BY sf.due_date
    """, (student_id,))
    
    columns = [column[0] for column in cursor.description]
    fees = []
    
    for row in cursor.fetchall():
        fee = dict(zip(columns, row))
        fee['balance'] = fee['amount'] - fee['discount_amount'] - fee['total_paid']
        
        # Get installments if this fee has an installment plan
        if fee.get('installment_plan_id'):
            installments = StudentInstallment.get_student_installments(fee['student_fee_id'])
            fee['installments'] = installments
        else:
            fee['installments'] = []
        
        # Get individual payments for this fee
        cursor.execute("""
        SELECT * FROM FeePayments 
        WHERE student_fee_id = ? 
        ORDER BY payment_date DESC
        """, (fee['student_fee_id'],))
        
        payment_columns = [column[0] for column in cursor.description]
        payments = []
        for p_row in cursor.fetchall():
            payments.append(dict(zip(payment_columns, p_row)))
        
        fee['payments'] = payments
        fee['payment_count'] = len(payments)  # Add payment_count here
        fees.append(fee)
    
    # Get ALL fee categories
    cursor.execute("SELECT * FROM FeeCategories ORDER BY category_name_ar")
    category_columns = [column[0] for column in cursor.description]
    categories = []
    for row in cursor.fetchall():
        categories.append(dict(zip(category_columns, row)))
    
    today = date.today().isoformat()
    cursor.close()
    
    return render_template('finance/student_fees_detail.html',
                         student=student,
                         fees=fees,
                         categories=categories,
                         today=today)
@finance_bp.route('/student/<int:student_id>/payments')
def student_payments(student_id):
    """Get payment history for a student"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("""
    SELECT 
        fp.payment_id,
        CONVERT(varchar, fp.payment_date, 23) as payment_date,
        fp.amount_paid,
        fp.receipt_number,
        fp.payment_method,
        fp.notes,
        fc.category_name_ar as fee_category,
        u.full_name_ar as received_by
    FROM FeePayments fp
    JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
    JOIN FeeCategories fc ON sf.fee_category_id = fc.fee_category_id
    JOIN Users u ON fp.received_by = u.user_id
    WHERE sf.student_id = ?
    ORDER BY fp.payment_date DESC, fp.payment_id DESC
    """, (student_id,))
    
    payments = []
    for row in cursor.fetchall():
        payments.append({
            'payment_id': row[0],
            'payment_date': row[1],
            'amount_paid': float(row[2]),
            'receipt_number': row[3],
            'payment_method': row[4],
            'notes': row[5] or '',
            'fee_category': row[6],
            'received_by': row[7]
        })
    
    cursor.close()
    return jsonify(payments)

# ==================== EXPENSES ====================
@role_required('admin', 'accountant')
@finance_bp.route('/expenses')
def expenses():
    """List all expenses"""
    
    from datetime import datetime, timedelta
    
    db = get_db()
    cursor = db.cursor()
    
    # Get all expenses
    cursor.execute("""
    SELECT e.*, u.full_name_ar as entered_by_name
    FROM Expenses e
    LEFT JOIN Users u ON e.entered_by = u.user_id
    ORDER BY e.expense_date DESC, e.created_at DESC
    """)
    
    columns = [column[0] for column in cursor.description]
    expenses = []
    total_amount = 0
    
    for row in cursor.fetchall():
        expense = dict(zip(columns, row))
        expenses.append(expense)
        total_amount += expense['amount']
    
    # Calculate statistics
    today = datetime.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    today_amount = sum(e['amount'] for e in expenses if e['expense_date'] == today.isoformat())
    today_count = sum(1 for e in expenses if e['expense_date'] == today.isoformat())
    
    week_amount = sum(e['amount'] for e in expenses if e['expense_date'] >= week_ago.isoformat())
    week_count = sum(1 for e in expenses if e['expense_date'] >= week_ago.isoformat())
    
    month_amount = sum(e['amount'] for e in expenses if e['expense_date'] >= month_ago.isoformat())
    month_count = sum(1 for e in expenses if e['expense_date'] >= month_ago.isoformat())
    
    cursor.close()
    
    return render_template('finance/expenses.html',
                         expenses=expenses,
                         total_amount=f"{total_amount:,.0f}",
                         total_count=len(expenses),
                         today_amount=f"{today_amount:,.0f}",
                         today_count=today_count,
                         week_amount=f"{week_amount:,.0f}",
                         week_count=week_count,
                         month_amount=f"{month_amount:,.0f}",
                         month_count=month_count)

@finance_bp.route('/expenses/add', methods=['GET', 'POST'])
@role_required('admin', 'accountant')
def add_expense():
    """Add new expense"""
    
    from datetime import date
    today = date.today().isoformat()
    
    if request.method == 'POST':
        try:
            db = get_db()
            cursor = db.cursor()
            
            cursor.execute("""
            INSERT INTO Expenses (expense_date, expense_category, description, amount,
                                payment_method, reference_number, entered_by, notes)
            OUTPUT INSERTED.expense_id
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                request.form['expense_date'],
                request.form['expense_category'],
                request.form['description'],
                float(request.form['amount']),
                request.form['payment_method'],
                request.form.get('reference_number', ''),
                session['user_id'],
                request.form.get('notes', '')
            ))
            
            expense_id = cursor.fetchone()[0]
            db.commit()
            cursor.close()
            
            flash('Expense added successfully', 'success')
            return redirect(url_for('finance.expenses'))
            
        except Exception as e:
            flash(f'Error adding expense: {str(e)}', 'danger')
            db.rollback()
    
    return render_template('finance/add_expense.html', today=today)

@finance_bp.route('/receipt/<int:payment_id>/print-pdf')
def print_receipt_pdf(payment_id):
    """Generate PDF receipt"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    from weasyprint import HTML
    from flask import make_response
    
    language = session.get('language', 'ar')
    direction = 'rtl' if language == 'ar' else 'ltr'
    
    db = get_db()
    cursor = db.cursor()
    
    # Fetch receipt data (same as before)
    cursor.execute("""
    SELECT 
        fp.receipt_number,
        CONVERT(varchar, fp.payment_date, 23) as payment_date,
        fp.amount_paid,
        fp.payment_method,
        fp.notes,
        s.student_number,
        CONCAT(s.first_name_ar, ' ', s.last_name_ar) as student_name_ar,
        CONCAT(s.first_name_en, ' ', s.last_name_en) as student_name_en,
        c.class_name_ar,
        c.class_name_en,
        ay.year_name_ar as academic_year_ar,
        ay.year_name_en as academic_year_en,
        u.full_name_ar as accountant_name
    FROM FeePayments fp
    JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
    JOIN Students s ON sf.student_id = s.student_id
    JOIN Classes c ON s.current_class_id = c.class_id
    JOIN AcademicYears ay ON sf.academic_year_id = ay.year_id
    JOIN Users u ON fp.received_by = u.user_id
    WHERE fp.payment_id = ?
    """, (payment_id,))
    
    data = cursor.fetchone()
    cursor.close()
    
    # Prepare template data
    template_data = {
        'language': language,
        'direction': direction,
        'school_logo': url_for('static', filename='images/school_logo.png', _external=True),
        'school_name': school_settings.get('school_name_ar' if language == 'ar' else 'school_name_en'),
        'school_address': school_settings.get('school_address'),
        'school_phone': school_settings.get('school_phone'),
        'school_email': school_settings.get('school_email'),
        'receipt_number': data[0],
        'payment_date': data[1],
        'total_amount': float(data[2]),
        'payment_method': data[3],
        'notes': data[4] or '',
        'student_id': data[5],
        'student_name': data[6] if language == 'ar' else data[7],
        'class_name': data[8] if language == 'ar' else data[9],
        'academic_year': data[10] if language == 'ar' else data[11],
        'accountant_name': data[12],
        'currency': 'SDG',
        'payment_status': 'paid',
        'fee_items': [
            {
                'description': fee_category['category_name_ar' if language == 'ar' else 'category_name_en'],
                'amount': float(data[2])
            }
        ],
        'current_year': datetime.now().year
    }
    
    # Render HTML
    html = render_template('finance/payment_receipt.html', **template_data)
    
    # Generate PDF
    pdf = HTML(string=html).write_pdf()
    
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=receipt_{data[0]}.pdf'
    
    return response
@finance_bp.route('/receipt/<int:payment_id>/pdf')
def download_receipt_pdf(payment_id):
    """Download payment receipt as PDF"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    # Get user's language preference
    language = session.get('language', 'ar')
    
    db = get_db()
    cursor = db.cursor()
    
    # Get payment details with student information
    cursor.execute("""
    SELECT 
        fp.payment_id,
        CONVERT(varchar, fp.payment_date, 23) as payment_date,
        fp.amount_paid,
        fp.receipt_number,
        fp.payment_method,
        ISNULL(fp.notes, '') as notes,
        sf.student_id,
        sf.fee_category_id,
        s.first_name_ar, 
        s.last_name_ar, 
        s.first_name_en, 
        s.last_name_en,
        s.student_number, 
        c.class_name_ar, 
        c.class_name_en
    FROM FeePayments fp
    JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
    JOIN Students s ON sf.student_id = s.student_id
    JOIN Classes c ON s.current_class_id = c.class_id
    WHERE fp.payment_id = ?
    """, (payment_id,))
    
    payment_row = cursor.fetchone()
    if not payment_row:
        flash('Payment not found', 'danger')
        return redirect(url_for('finance.dashboard'))
    
    # Create payment dictionary
    payment = {
        'payment_id': payment_row[0],
        'payment_date': payment_row[1],
        'amount_paid': float(payment_row[2]),
        'receipt_number': payment_row[3],
        'payment_method': payment_row[4],
        'notes': payment_row[5],
        'student_id': payment_row[6],
        'fee_category_id': payment_row[7]
    }
    
    # Create student dictionary with all required fields
    student = {
        'student_id': payment_row[6],
        'first_name_ar': payment_row[8],
        'last_name_ar': payment_row[9],
        'first_name_en': payment_row[10],
        'last_name_en': payment_row[11],
        'student_number': payment_row[12],
        'class_name_ar': payment_row[13],
        'class_name_en': payment_row[14]
    }
    
    # Get fee category
    cursor.execute("SELECT * FROM FeeCategories WHERE fee_category_id = ?", 
                  (payment['fee_category_id'],))
    fee_columns = [column[0] for column in cursor.description]
    fee_category = dict(zip(fee_columns, cursor.fetchone()))
    
    # Get school settings
    cursor.execute("SELECT setting_key, setting_value FROM SystemSettings")
    school_settings = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Get user info
    cursor.execute("SELECT full_name_ar FROM Users WHERE user_id = ?", (session['user_id'],))
    user_row = cursor.fetchone()
    user = {'full_name_ar': user_row[0] if user_row else session.get('full_name', 'User')}
    
    cursor.close()
    
    from utils.pdf_receipt import PDFReceiptGenerator
    pdf_buffer = PDFReceiptGenerator.generate_payment_receipt(
        payment, student, fee_category, school_settings, user, language
    )
    
    return send_file(
        pdf_buffer,
        download_name=f"receipt_{payment['receipt_number']}.pdf",
        as_attachment=True,
        mimetype='application/pdf'
    )
@finance_bp.route('/reports/export/<string:report_type>/<string:format>')
@role_required('admin', 'accountant')
def export_report(report_type, format):
    """Export reports in various formats"""
    
    db = get_db()
    cursor = db.cursor()
    
    filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d')}"
    
    if report_type == 'students':
        # Get student data
        cursor.execute("""
        SELECT s.*, c.class_name_ar
        FROM Students s
        LEFT JOIN Classes c ON s.current_class_id = c.class_id
        WHERE s.status = 'active'
        ORDER BY s.student_id
        """)
        
        columns = [column[0] for column in cursor.description]
        students = []
        for row in cursor.fetchall():
            students.append(dict(zip(columns, row)))
        
        if format == 'excel':
            output = ExcelReportGenerator.generate_student_report(students)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename += '.xlsx'
        elif format == 'pdf':
            output = PDFReportGenerator.generate_student_report(students)
            mimetype = 'application/pdf'
            filename += '.pdf'
        else:
            flash('Invalid format', 'danger')
            return redirect(url_for('finance.reports'))
    
    elif report_type == 'fees':
        # Get fees data
        cursor.execute("""
        SELECT 
            sf.*,
            s.student_number,
            CONCAT(s.first_name_ar, ' ', s.last_name_ar) as student_name,
            c.class_name_ar as class_name,
            fc.category_name_ar as fee_type,
            ISNULL(SUM(fp.amount_paid), 0) as paid
        FROM StudentFees sf
        JOIN Students s ON sf.student_id = s.student_id
        JOIN Classes c ON s.current_class_id = c.class_id
        JOIN FeeCategories fc ON sf.fee_category_id = fc.fee_category_id
        LEFT JOIN FeePayments fp ON sf.student_fee_id = fp.student_fee_id
        GROUP BY sf.student_fee_id, s.student_number, s.first_name_ar, s.last_name_ar,
                 c.class_name_ar, fc.category_name_ar, sf.amount, 
                 sf.discount_amount, sf.due_date, sf.status
        """)
        
        columns = [column[0] for column in cursor.description]
        fees = []
        for row in cursor.fetchall():
            fee = dict(zip(columns, row))
            fee['balance'] = fee['amount'] - fee['discount_amount'] - fee['paid']
            fees.append(fee)
        
        if format == 'excel':
            output = ExcelReportGenerator.generate_fees_report(fees)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename += '.xlsx'
        elif format == 'pdf':
            output = PDFReportGenerator.generate_fees_report(fees)
            mimetype = 'application/pdf'
            filename += '.pdf'
        else:
            flash('Invalid format', 'danger')
            return redirect(url_for('finance.reports'))
    
    elif report_type == 'expenses':
        # Get expenses data
        cursor.execute("SELECT * FROM Expenses ORDER BY expense_date DESC")
        columns = [column[0] for column in cursor.description]
        expenses = []
        for row in cursor.fetchall():
            expenses.append(dict(zip(columns, row)))
        
        if format == 'excel':
            output = ExcelReportGenerator.generate_expenses_report(expenses)
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename += '.xlsx'
        else:
            flash('PDF export for expenses coming soon', 'info')
            return redirect(url_for('finance.expenses'))
    
    else:
        flash('Invalid report type', 'danger')
        return redirect(url_for('finance.reports'))
    
    cursor.close()
    
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype=mimetype
    )
@finance_bp.route('/expenses/<int:expense_id>/data')
def get_expense_data(expense_id):
    """Get expense data for modal"""
    if 'user_id' not in session:
        return jsonify({'error': 'Access denied'}), 403
    
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("""
    SELECT e.*, u.full_name_ar as entered_by_name
    FROM Expenses e
    LEFT JOIN Users u ON e.entered_by = u.user_id
    WHERE e.expense_id = ?
    """, (expense_id,))
    
    columns = [column[0] for column in cursor.description]
    row = cursor.fetchone()
    cursor.close()
    
    if row:
        expense = dict(zip(columns, row))
        return jsonify(expense)
    
    return jsonify({'error': 'Expense not found'}), 404

@finance_bp.route('/expenses/<int:expense_id>/edit', methods=['GET', 'POST'])
@role_required('admin')
def edit_expense(expense_id):
    """Edit expense"""
    
    db = get_db()
    cursor = db.cursor()
    
    if request.method == 'POST':
        try:
            cursor.execute("""
            UPDATE Expenses
            SET expense_date = ?, expense_category = ?, description = ?,
                amount = ?, payment_method = ?, reference_number = ?, notes = ?
            WHERE expense_id = ?
            """, (
                request.form['expense_date'],
                request.form['expense_category'],
                request.form['description'],
                float(request.form['amount']),
                request.form['payment_method'],
                request.form.get('reference_number', ''),
                request.form.get('notes', ''),
                expense_id
            ))
            
            db.commit()
            cursor.close()
            flash('Expense updated successfully', 'success')
            return redirect(url_for('finance.expenses'))
            
        except Exception as e:
            flash(f'Error updating expense: {str(e)}', 'danger')
            db.rollback()
    
    # GET request - show edit form
    cursor.execute("SELECT * FROM Expenses WHERE expense_id = ?", (expense_id,))
    columns = [column[0] for column in cursor.description]
    expense = dict(zip(columns, cursor.fetchone()))
    cursor.close()
    
    return render_template('finance/edit_expense.html', expense=expense)

@finance_bp.route('/expenses/<int:expense_id>/delete', methods=['POST'])
@role_required('admin')
def delete_expense(expense_id):
    """Delete expense"""
    
    try:
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("DELETE FROM Expenses WHERE expense_id = ?", (expense_id,))
        db.commit()
        cursor.close()
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# ==================== REPORTS ====================
@finance_bp.route('/reports')
@role_required('admin', 'accountant')
def reports():
    """Financial reports"""

    
    from datetime import datetime, timedelta
    
    db = get_db()
    cursor = db.cursor()
    
    # Get date range from request or default to current month
    period = request.args.get('period', 'month')
    today = datetime.now()
    
    if period == 'today':
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
        prev_start = (today - timedelta(days=1)).strftime('%Y-%m-%d')
        prev_end = (today - timedelta(days=1)).strftime('%Y-%m-%d')
    elif period == 'week':
        start_date = (today - timedelta(days=7)).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
        prev_start = (today - timedelta(days=14)).strftime('%Y-%m-%d')
        prev_end = (today - timedelta(days=7)).strftime('%Y-%m-%d')
    elif period == 'month':
        start_date = today.strftime('%Y-%m-01')
        end_date = today.strftime('%Y-%m-%d')
        # Previous month
        if today.month == 1:
            prev_start = datetime(today.year-1, 12, 1).strftime('%Y-%m-01')
            prev_end = datetime(today.year-1, 12, 31).strftime('%Y-%m-%d')
        else:
            prev_start = datetime(today.year, today.month-1, 1).strftime('%Y-%m-01')
            prev_end = datetime(today.year, today.month, 1).strftime('%Y-%m-%d')
    else:  # year
        start_date = today.strftime('%Y-01-01')
        end_date = today.strftime('%Y-%m-%d')
        prev_start = datetime(today.year-1, 1, 1).strftime('%Y-01-01')
        prev_end = datetime(today.year-1, 12, 31).strftime('%Y-%m-%d')
    
    # Get real income data (payments)
    cursor.execute("""
    SELECT 
        CONVERT(varchar, fp.payment_date, 23) as date,
        CONCAT(s.first_name_ar, ' ', s.last_name_ar) as description,
        fc.category_name_ar as category,
        fp.amount_paid as amount,
        fp.receipt_number as reference,
        'income' as type
    FROM FeePayments fp
    JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
    JOIN Students s ON sf.student_id = s.student_id
    JOIN FeeCategories fc ON sf.fee_category_id = fc.fee_category_id
    WHERE fp.payment_date BETWEEN ? AND ?
    ORDER BY fp.payment_date DESC
    """, (start_date, end_date))
    
    income_data = []
    for row in cursor.fetchall():
        income_data.append({
            'date': row[0],
            'description': row[1],
            'category': row[2],
            'amount': float(row[3]),
            'reference': row[4],
            'type': 'income'
        })
    
    # Get real expense data
    cursor.execute("""
    SELECT 
        CONVERT(varchar, expense_date, 23) as date,
        description,
        expense_category as category,
        amount,
        reference_number as reference,
        'expense' as type
    FROM Expenses
    WHERE expense_date BETWEEN ? AND ?
    ORDER BY expense_date DESC
    """, (start_date, end_date))
    
    expense_data = []
    for row in cursor.fetchall():
        expense_data.append({
            'date': row[0],
            'description': row[1],
            'category': row[2],
            'amount': float(row[3]),
            'reference': row[4],
            'type': 'expense'
        })
    
    # Combine and sort all transactions
    all_transactions = income_data + expense_data
    all_transactions.sort(key=lambda x: x['date'], reverse=True)
    
    # Calculate totals
    total_income = sum(t['amount'] for t in income_data)
    total_expenses = sum(t['amount'] for t in expense_data)
    net_income = total_income - total_expenses
    
    # Get summary by category
    cursor.execute("""
    SELECT 
        fc.category_name_ar,
        COUNT(*) as count,
        SUM(fp.amount_paid) as total
    FROM FeePayments fp
    JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
    JOIN FeeCategories fc ON sf.fee_category_id = fc.fee_category_id
    WHERE fp.payment_date BETWEEN ? AND ?
    GROUP BY fc.category_name_ar
    ORDER BY total DESC
    """, (start_date, end_date))
    
    income_by_category = []
    for row in cursor.fetchall():
        income_by_category.append({
            'category': row[0],
            'count': row[1],
            'total': float(row[2])
        })
    
    cursor.execute("""
    SELECT 
        expense_category,
        COUNT(*) as count,
        SUM(amount) as total
    FROM Expenses
    WHERE expense_date BETWEEN ? AND ?
    GROUP BY expense_category
    ORDER BY total DESC
    """, (start_date, end_date))
    
    expenses_by_category = []
    for row in cursor.fetchall():
        expenses_by_category.append({
            'category': row[0],
            'count': row[1],
            'total': float(row[2])
        })
    
    # Get chart data for the last 7 days
    chart_labels = []
    income_chart_data = []
    expense_chart_data = []
    
    for i in range(6, -1, -1):
        date = (today - timedelta(days=i)).strftime('%Y-%m-%d')
        chart_labels.append((today - timedelta(days=i)).strftime('%a'))
        
        cursor.execute("""
        SELECT ISNULL(SUM(amount_paid), 0) FROM FeePayments 
        WHERE payment_date = ?
        """, (date,))
        income_chart_data.append(float(cursor.fetchone()[0]))
        
        cursor.execute("""
        SELECT ISNULL(SUM(amount), 0) FROM Expenses 
        WHERE expense_date = ?
        """, (date,))
        expense_chart_data.append(float(cursor.fetchone()[0]))
    
    cursor.close()
    
    return render_template('finance/reports.html',
                         period=period,
                         start_date=start_date,
                         end_date=end_date,
                         total_income=total_income,
                         total_expenses=total_expenses,
                         net_income=net_income,
                         income_count=len(income_data),
                         expense_count=len(expense_data),
                         transactions=all_transactions[:50],  # Last 50 transactions
                         income_by_category=income_by_category,
                         expenses_by_category=expenses_by_category,
                         chart_labels=chart_labels,
                         income_chart_data=income_chart_data,
                         expense_chart_data=expense_chart_data)

@finance_bp.route('/reports/income-statement')
def income_statement():
    """Generate income statement"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    return render_template('finance/income_statement.html')

@finance_bp.route('/reports/export/<string:format>')
@role_required('admin', 'accountant')
def export_report_data(format):
    """Export report data to file"""
    
    from datetime import datetime, timedelta
    import csv
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    # Get date range
    period = request.args.get('period', 'month')
    today = datetime.now()
    
    if period == 'today':
        start_date = today.strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
        period_text = 'Today' if session.get('language') != 'ar' else 'اليوم'
    elif period == 'week':
        start_date = (today - timedelta(days=7)).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
        period_text = 'This Week' if session.get('language') != 'ar' else 'هذا الأسبوع'
    elif period == 'month':
        start_date = today.strftime('%Y-%m-01')
        end_date = today.strftime('%Y-%m-%d')
        period_text = 'This Month' if session.get('language') != 'ar' else 'هذا الشهر'
    else:  # year
        start_date = today.strftime('%Y-01-01')
        end_date = today.strftime('%Y-%m-%d')
        period_text = 'This Year' if session.get('language') != 'ar' else 'هذه السنة'
    
    # Override with custom dates if provided
    if request.args.get('start_date'):
        start_date = request.args['start_date']
    if request.args.get('end_date'):
        end_date = request.args['end_date']
    
    db = get_db()
    cursor = db.cursor()
    
    # Get income summary
    cursor.execute("""
    SELECT 
        ISNULL(SUM(amount_paid), 0) as total,
        COUNT(*) as count
    FROM FeePayments
    WHERE payment_date BETWEEN ? AND ?
    """, (start_date, end_date))
    row = cursor.fetchone()
    total_income = float(row[0]) if row else 0
    income_count = row[1] if row else 0
    
    # Get expense summary
    cursor.execute("""
    SELECT 
        ISNULL(SUM(amount), 0) as total,
        COUNT(*) as count
    FROM Expenses
    WHERE expense_date BETWEEN ? AND ?
    """, (start_date, end_date))
    row = cursor.fetchone()
    total_expenses = float(row[0]) if row else 0
    expense_count = row[1] if row else 0
    
    # Get all transactions for the report
    cursor.execute("""
    SELECT 
        CONVERT(varchar, fp.payment_date, 23) as date,
        CONCAT(s.first_name_ar, ' ', s.last_name_ar) as description,
        fc.category_name_ar as category,
        fp.amount_paid as amount,
        fp.receipt_number as reference,
        'إيراد' as type
    FROM FeePayments fp
    JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
    JOIN Students s ON sf.student_id = s.student_id
    JOIN FeeCategories fc ON sf.fee_category_id = fc.fee_category_id
    WHERE fp.payment_date BETWEEN ? AND ?
    UNION ALL
    SELECT 
        CONVERT(varchar, expense_date, 23) as date,
        description,
        expense_category as category,
        amount,
        ISNULL(reference_number, '') as reference,
        'مصروف' as type
    FROM Expenses
    WHERE expense_date BETWEEN ? AND ?
    ORDER BY date DESC
    """, (start_date, end_date, start_date, end_date))
    
    transactions = cursor.fetchall()
    cursor.close()
    
    # Generate filename
    filename = f"financial_report_{start_date}_to_{end_date}"
    
    # Handle different export formats
    if format == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow(['التاريخ', 'الوصف', 'الفئة', 'المبلغ', 'المرجع', 'النوع'])
        
        # Write summary
        writer.writerow([])
        writer.writerow(['ملخص', '', '', '', '', ''])
        writer.writerow(['إجمالي الإيرادات', '', '', f"{total_income:.2f}", f"عدد: {income_count}", ''])
        writer.writerow(['إجمالي المصروفات', '', '', f"{total_expenses:.2f}", f"عدد: {expense_count}", ''])
        writer.writerow(['صافي الدخل', '', '', f"{total_income - total_expenses:.2f}", '', ''])
        writer.writerow([])
        writer.writerow(['التفاصيل', '', '', '', '', ''])
        
        # Write transactions
        for t in transactions:
            writer.writerow(t)
        
        output.seek(0)
        
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"{filename}.csv"
        )
    
    elif format == 'excel':
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Prepare data for DataFrame
            data = []
            for t in transactions:
                data.append({
                    'التاريخ': t[0],
                    'الوصف': t[1],
                    'الفئة': t[2],
                    'المبلغ': float(t[3]),
                    'المرجع': t[4],
                    'النوع': t[5]
                })
            
            df = pd.DataFrame(data)
            
            # Create Excel file
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write summary sheet
                summary_df = pd.DataFrame({
                    'البند': ['إجمالي الإيرادات', 'إجمالي المصروفات', 'صافي الدخل', 'فترة التقرير'],
                    'القيمة': [f"{total_income:.2f} SDG", f"{total_expenses:.2f} SDG", 
                              f"{total_income - total_expenses:.2f} SDG", 
                              f"{start_date} إلى {end_date}"]
                })
                summary_df.to_excel(writer, sheet_name='ملخص', index=False)
                
                # Write transactions sheet
                df.to_excel(writer, sheet_name='المعاملات', index=False)
                
                # Format the Excel file
                workbook = writer.book
                
                # Format summary sheet
                summary_sheet = workbook['ملخص']
                for row in summary_sheet.iter_rows(min_row=2, max_row=4, min_col=2, max_col=2):
                    for cell in row:
                        if 'SDG' in str(cell.value):
                            cell.font = Font(color='28A745' if 'إيرادات' in str(cell.value) else 'DC3545')
                
                # Format transactions sheet
                trans_sheet = workbook['المعاملات']
                # Bold headers
                for cell in trans_sheet[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='875A7B', end_color='875A7B', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF')
                
                # Adjust column widths
                for column in trans_sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    trans_sheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f"{filename}.xlsx"
            )
            
        except ImportError:
            flash('Please install pandas and openpyxl: pip install pandas openpyxl', 'danger')
            return redirect(url_for('finance.reports'))
    
    elif format == 'pdf':
        # Create PDF with Arabic font support
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=18)
        story = []
        
        # Register Arabic font if available
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # Try to register an Arabic font
            font_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'fonts', 'Amiri-Regular.ttf')
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('Arabic', font_path))
                arabic_font = 'Arabic'
            else:
                # Fallback to Helvetica (will show squares but at least it works)
                arabic_font = 'Helvetica'
        except:
            arabic_font = 'Helvetica'
        
        # Create styles
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=arabic_font,
            fontSize=18,
            textColor=colors.HexColor('#875A7B'),
            alignment=1,  # Center alignment
            spaceAfter=20,
            encoding='utf-8'
        )
        
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontName=arabic_font,
            fontSize=10,
            alignment=1,  # Center
            encoding='utf-8'
        )
        
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontName=arabic_font,
            fontSize=12,
            textColor=colors.whitesmoke,
            alignment=1,
            encoding='utf-8'
        )
        
        # Title
        story.append(Paragraph("التقرير المالي", title_style))
        story.append(Paragraph(f"الفترة: {start_date} إلى {end_date}", normal_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Summary Table
        summary_data = [
            [Paragraph('البيان', header_style), Paragraph('المبلغ', header_style)],
            [Paragraph('إجمالي الإيرادات', normal_style), Paragraph(f"{total_income:,.2f} SDG", normal_style)],
            [Paragraph('إجمالي المصروفات', normal_style), Paragraph(f"{total_expenses:,.2f} SDG", normal_style)],
            [Paragraph('صافي الدخل', normal_style), Paragraph(f"{total_income - total_expenses:,.2f} SDG", normal_style)],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#875A7B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), arabic_font),
            ('FONTNAME', (0, 1), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Transactions Table
        trans_data = [[
            Paragraph('التاريخ', header_style),
            Paragraph('الوصف', header_style),
            Paragraph('الفئة', header_style),
            Paragraph('المبلغ', header_style),
            Paragraph('المرجع', header_style)
        ]]
        
        for t in transactions[:50]:  # Limit to 50 transactions for PDF
            amount = float(t[3])
            amount_color = colors.HexColor('#28A745') if t[5] == 'إيراد' else colors.HexColor('#DC3545')
            
            trans_data.append([
                Paragraph(str(t[0]), normal_style),
                Paragraph(str(t[1])[:30], normal_style),
                Paragraph(str(t[2]), normal_style),
                Paragraph(f"{amount:,.2f} SDG", normal_style),
                Paragraph(str(t[4] or '---'), normal_style)
            ])
        
        trans_table = Table(trans_data, colWidths=[1.2*inch, 2.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        
        # Build table style
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#17A2B8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), arabic_font),
            ('FONTNAME', (0, 1), (-1, -1), arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]
        
        # Add row colors
        for i in range(1, len(trans_data)):
            if i % 2 == 0:
                table_style.append(('BACKGROUND', (0, i), (-1, i), colors.lightgrey))
        
        trans_table.setStyle(TableStyle(table_style))
        story.append(trans_table)
        
        # Footer with generation info
        story.append(Spacer(1, 0.2*inch))
        footer_text = f"تم إنشاء هذا التقرير في {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        story.append(Paragraph(footer_text, normal_style))
        
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{filename}.pdf"
        )
    
    else:
        flash('Invalid export format', 'danger')
        return redirect(url_for('finance.reports'))
# ==================== RECEIPTS ====================
@finance_bp.route('/receipt/<int:payment_id>/print')
def print_receipt(payment_id):
    """Print payment receipt (opens in browser for printing)"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    language = session.get('language', 'ar')
    
    db = get_db()
    cursor = db.cursor()
    
    # Get payment details with all necessary information
    cursor.execute("""
    SELECT 
        fp.payment_id,
        CONVERT(varchar, fp.payment_date, 23) as payment_date,
        fp.amount_paid,
        fp.receipt_number,
        fp.payment_method,
        ISNULL(fp.notes, '') as notes,
        sf.student_id,
        sf.fee_category_id,
        s.first_name_ar, 
        s.last_name_ar, 
        s.first_name_en, 
        s.last_name_en,
        s.student_number, 
        c.class_name_ar, 
        c.class_name_en,
        ay.year_name_ar as academic_year_ar,
        ay.year_name_en as academic_year_en,
        fc.category_name_ar,
        fc.category_name_en
    FROM FeePayments fp
    JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
    JOIN Students s ON sf.student_id = s.student_id
    JOIN Classes c ON s.current_class_id = c.class_id
    JOIN AcademicYears ay ON sf.academic_year_id = ay.year_id
    JOIN FeeCategories fc ON sf.fee_category_id = fc.fee_category_id
    WHERE fp.payment_id = ?
    """, (payment_id,))
    
    payment_row = cursor.fetchone()
    if not payment_row:
        flash('Payment not found', 'danger')
        return redirect(url_for('finance.dashboard'))
    
    # Create payment dictionary
    payment = {
        'payment_id': payment_row[0],
        'payment_date': payment_row[1],
        'amount_paid': float(payment_row[2]),
        'receipt_number': payment_row[3],
        'payment_method': payment_row[4],
        'notes': payment_row[5],
        'student_id': payment_row[6],
        'fee_category_id': payment_row[7]
    }
    
    # Create student dictionary
    student = {
        'student_id': payment_row[6],
        'first_name_ar': payment_row[8],
        'last_name_ar': payment_row[9],
        'first_name_en': payment_row[10],
        'last_name_en': payment_row[11],
        'student_number': payment_row[12],
        'class_name_ar': payment_row[13],
        'class_name_en': payment_row[14]
    }
    
    # Get academic year based on language
    academic_year = payment_row[15] if language == 'ar' else payment_row[16]
    
    # Create fee category dictionary
    fee_category = {
        'category_name_ar': payment_row[17],
        'category_name_en': payment_row[18]
    }
    
    # Get school settings
    cursor.execute("SELECT setting_key, setting_value FROM SystemSettings")
    school_settings = {row[0]: row[1] for row in cursor.fetchall()}
    school_settings['academic_year'] = academic_year
    
    # Get user info
    cursor.execute("SELECT full_name_ar FROM Users WHERE user_id = ?", (session['user_id'],))
    user_row = cursor.fetchone()
    user = {'full_name_ar': user_row[0] if user_row else session.get('full_name', 'User')}
    
    cursor.close()
    
    from utils.pdf_receipt import PDFReceiptGenerator
    pdf_buffer = PDFReceiptGenerator.generate_payment_receipt(
        payment, student, fee_category, school_settings, user, language
    )
    
    return send_file(
        pdf_buffer,
        download_name=f"receipt_{payment['receipt_number']}.pdf",
        as_attachment=False,
        mimetype='application/pdf'
    )
@finance_bp.route('/fee/<int:fee_id>/pay', methods=['POST'])
@role_required('admin', 'accountant')
def make_payment(fee_id):
    """Make payment for a fee"""
    
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Get the student_fee_id from the fee_id (if fee_id is actually the student_fee_id)
        # In your form, you're likely passing the student_fee_id as fee_id
        
        data = {
            'student_fee_id': fee_id,  # This should be the student_fee_id
            'payment_date': request.form['payment_date'],
            'amount_paid': float(request.form['amount_paid']),
            'payment_method': request.form['payment_method'],
            'received_by': session['user_id'],
            'notes': request.form.get('notes', '')
        }
        
        # Generate receipt number
        year = datetime.now().strftime('%Y')
        cursor.execute("""
            SELECT ISNULL(MAX(CAST(RIGHT(receipt_number, 6) AS INT)), 0) + 1 
            FROM FeePayments 
            WHERE receipt_number LIKE ?
        """, (f"RCP-{year}%",))
        
        next_num = cursor.fetchone()[0]
        receipt_number = f"RCP-{year}-{str(next_num).zfill(6)}"
        
        # Insert payment
        cursor.execute("""
        INSERT INTO FeePayments (student_fee_id, payment_date, amount_paid, 
                                payment_method, receipt_number, received_by, notes)
        OUTPUT INSERTED.payment_id
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            data['student_fee_id'],
            data['payment_date'],
            data['amount_paid'],
            data['payment_method'],
            receipt_number,
            data['received_by'],
            data['notes']
        ))
        
        payment_id = cursor.fetchone()[0]
        
        # Update fee status
        cursor.execute("""
            SELECT SUM(amount_paid) as total_paid, sf.amount, sf.discount_amount
            FROM FeePayments fp
            JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
            WHERE fp.student_fee_id = ?
            GROUP BY sf.amount, sf.discount_amount
        """, (fee_id,))
        
        result = cursor.fetchone()
        if result:
            total_paid = result[0]
            amount = result[1]
            discount = result[2]
            net_amount = amount - discount
            
            if total_paid >= net_amount:
                status = 'paid'
            elif total_paid > 0:
                status = 'partial'
            else:
                status = 'pending'
            
            cursor.execute("UPDATE StudentFees SET status = ? WHERE student_fee_id = ?", 
                         (status, fee_id))
        
        db.commit()
        cursor.close()
        
        flash(f'Payment recorded successfully. Receipt: {receipt_number}', 'success')
        
    except Exception as e:
        flash(f'Error recording payment: {str(e)}', 'danger')
        db.rollback()
    
    # Get student_id from fee
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT student_id FROM StudentFees WHERE student_fee_id = ?", (fee_id,))
    result = cursor.fetchone()
    student_id = result[0] if result else None
    cursor.close()
    
    if student_id:
        return redirect(url_for('finance.student_fees_detail', student_id=student_id))
    else:
        return redirect(url_for('finance.student_fees'))
@finance_bp.route('/preview-receipt')
def preview_receipt_with_settings():
    """Preview receipt with current logo settings"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    from utils.pdf_receipt import PDFReceiptGenerator, LOGO_CONFIG
    import json
    
    # Get logo settings from database
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT setting_value FROM SystemSettings WHERE setting_key = 'logo_config'")
    result = cursor.fetchone()
    
    logo_config = LOGO_CONFIG.copy()
    if result and result[0]:
        try:
            saved_config = json.loads(result[0])
            logo_config.update(saved_config)
        except:
            pass
    
    # Override with query parameters for preview
    if request.args.get('width'):
        logo_config['width'] = float(request.args.get('width')) * 72
    if request.args.get('height'):
        logo_config['height'] = float(request.args.get('height')) * 72
    if request.args.get('alignment'):
        logo_config['alignment'] = request.args.get('alignment')
    if request.args.get('border'):
        logo_config['border'] = request.args.get('border').lower() == 'true'
    
    # Get a sample payment for preview (use the most recent payment)
    cursor.execute("""
    SELECT TOP 1 fp.*, sf.student_id, sf.fee_category_id,
           s.first_name_ar, s.last_name_ar, s.first_name_en, s.last_name_en,
           s.student_number, c.class_name_ar, c.class_name_en
    FROM FeePayments fp
    JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
    JOIN Students s ON sf.student_id = s.student_id
    JOIN Classes c ON s.current_class_id = c.class_id
    ORDER BY fp.payment_id DESC
    """)
    
    payment_row = cursor.fetchone()
    
    if not payment_row:
        flash('No payments found for preview', 'warning')
        return redirect(url_for('finance.dashboard'))
    
    columns = [column[0] for column in cursor.description]
    payment = dict(zip(columns, payment_row))
    
    # Get fee category
    cursor.execute("SELECT * FROM FeeCategories WHERE fee_category_id = ?", 
                  (payment['fee_category_id'],))
    fee_columns = [column[0] for column in cursor.description]
    fee_category = dict(zip(fee_columns, cursor.fetchone()))
    
    # Get school settings
    cursor.execute("SELECT setting_key, setting_value FROM SystemSettings")
    school_settings = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Get user info
    cursor.execute("SELECT full_name_ar FROM Users WHERE user_id = ?", (session['user_id'],))
    user = {'full_name_ar': cursor.fetchone()[0]}
    
    cursor.close()
    
    # Generate PDF with custom logo settings
    pdf_buffer = PDFReceiptGenerator.generate_payment_receipt(
        payment, payment, fee_category, school_settings, user, logo_config
    )
    
    return send_file(
        pdf_buffer,
        download_name="receipt_preview.pdf",
        as_attachment=False,
        mimetype='application/pdf'
    )
@finance_bp.route('/student/<int:student_id>/fees/add', methods=['POST'])
@role_required('admin', 'accountant')
def add_fee_to_student(student_id):
    """Add fee to student with installment option"""
    
    try:
        db = get_db()
        cursor = db.cursor()
        
        # Get current academic year
        cursor.execute("SELECT year_id FROM AcademicYears WHERE is_current = 1")
        academic_year = cursor.fetchone()
        if not academic_year:
            flash('No active academic year found', 'danger')
            return redirect(url_for('finance.student_fees_detail', student_id=student_id))
        
        installment_plan_id = int(request.form.get('installment_plan_id', 1))
        
        data = {
            'student_id': student_id,
            'fee_category_id': request.form['fee_category_id'],
            'academic_year_id': academic_year[0],
            'amount': float(request.form['amount']),
            'discount_amount': float(request.form.get('discount_amount', 0)),
            'due_date': request.form['due_date'],
            'notes': request.form.get('notes', ''),
            'installment_plan_id': installment_plan_id if installment_plan_id > 1 else None,
            'allow_partial_payment': 1 if installment_plan_id > 1 else 0
        }
        
        # Insert the fee
        cursor.execute("""
        INSERT INTO StudentFees (student_id, fee_category_id, academic_year_id, 
                                amount, discount_amount, due_date, status, notes,
                                installment_plan_id, allow_partial_payment)
        OUTPUT INSERTED.student_fee_id
        VALUES (?, ?, ?, ?, ?, ?, 'pending', ?, ?, ?)
        """, (
            data['student_id'],
            data['fee_category_id'],
            data['academic_year_id'],
            data['amount'],
            data['discount_amount'],
            data['due_date'],
            data['notes'],
            data['installment_plan_id'],
            data['allow_partial_payment']
        ))
        
        student_fee_id = cursor.fetchone()[0]
        
        # Create installments if more than one payment
        if installment_plan_id > 1:
            net_amount = data['amount'] - data['discount_amount']
            StudentInstallment.create_installments(student_fee_id, installment_plan_id, net_amount, data['due_date'])
        
        db.commit()
        cursor.close()
        
        flash('Fee added to student successfully', 'success')
        
    except Exception as e:
        flash(f'Error adding fee: {str(e)}', 'danger')
        db.rollback()
    
    return redirect(url_for('finance.student_fees_detail', student_id=student_id))

@finance_bp.route('/installment/pay', methods=['POST'])
@role_required('admin', 'accountant')
def pay_installment():
    """Pay an installment"""
    
    try:
        installment_id = request.form['installment_id']
        amount_paid = float(request.form['amount_paid'])
        payment_date = request.form['payment_date']
        payment_method = request.form['payment_method']
        notes = request.form.get('notes', '')
        
        result = StudentInstallment.make_installment_payment(
            installment_id, amount_paid, payment_date, payment_method, 
            session['user_id'], notes
        )
        
        if result:
            flash(f'Payment recorded successfully. Receipt: {result["receipt_number"]}', 'success')
        else:
            flash('Error recording payment', 'danger')
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(request.referrer or url_for('finance.dashboard'))

@finance_bp.route('/installment/<int:installment_id>/details')
def get_installment_details(installment_id):
    
    """Get installment details for AJAX"""
    if 'user_id' not in session:
        return jsonify({'error': 'Access denied'}), 403
    
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute("""
    SELECT si.*, sf.amount as total_fee, sf.discount_amount
    FROM StudentInstallments si
    JOIN StudentFees sf ON si.student_fee_id = sf.student_fee_id
    WHERE si.installment_id = ?
    """, (installment_id,))
    
    row = cursor.fetchone()
    cursor.close()
    
    if row:
        return jsonify({
            'installment_id': row[0],
            'amount': float(row[3]),
            'paid_amount': float(row[4]),
            'remaining': float(row[3] - row[4]),
            'due_date': row[5].strftime('%Y-%m-%d') if row[5] else None
        })
    
    return jsonify({'error': 'Installment not found'}), 404

@finance_bp.route('/fee-categories/<int:category_id>/collections')
@role_required('admin', 'accountant')
def fee_category_collections(category_id):
    """View collection statistics for a fee category"""
    
    db = get_db()
    cursor = db.cursor()
    
    # Get fee category details
    cursor.execute("SELECT * FROM FeeCategories WHERE fee_category_id = ?", (category_id,))
    category = cursor.fetchone()
    if not category:
        flash('Fee category not found', 'danger')
        return redirect(url_for('finance.fee_categories'))
    
    category_dict = {
        'fee_category_id': category[0],
        'category_name_ar': category[1],
        'category_name_en': category[2],
        'description': category[3],
        'amount': float(category[4]),
        'is_annual': category[5],
        'is_mandatory': category[6]
    }
    
    # Get collection statistics
    cursor.execute("""
    SELECT 
        COUNT(DISTINCT sf.student_id) as total_students,
        COUNT(sf.student_fee_id) as total_fees,
        SUM(sf.amount) as total_amount,
        SUM(sf.discount_amount) as total_discounts,
        SUM(fp.total_paid) as total_collected,
        SUM(sf.amount - sf.discount_amount - ISNULL(fp.total_paid, 0)) as total_pending
    FROM StudentFees sf
    LEFT JOIN (
        SELECT student_fee_id, SUM(amount_paid) as total_paid
        FROM FeePayments
        GROUP BY student_fee_id
    ) fp ON sf.student_fee_id = fp.student_fee_id
    WHERE sf.fee_category_id = ?
    """, (category_id,))
    
    stats_row = cursor.fetchone()
    stats = {
        'total_students': stats_row[0] or 0,
        'total_fees': stats_row[1] or 0,
        'total_amount': float(stats_row[2] or 0),
        'total_discounts': float(stats_row[3] or 0),
        'total_collected': float(stats_row[4] or 0),
        'total_pending': float(stats_row[5] or 0)
    }
    stats['collection_percentage'] = round((stats['total_collected'] / stats['total_amount'] * 100), 1) if stats['total_amount'] > 0 else 0
    
    # Get collection by class
    cursor.execute("""
    SELECT 
        c.class_id,
        c.class_name_ar,
        COUNT(DISTINCT s.student_id) as students,
        SUM(sf.amount) as total_amount,
        SUM(sf.discount_amount) as total_discounts,
        SUM(fp.total_paid) as total_collected
    FROM StudentFees sf
    JOIN Students s ON sf.student_id = s.student_id
    JOIN Classes c ON s.current_class_id = c.class_id
    LEFT JOIN (
        SELECT student_fee_id, SUM(amount_paid) as total_paid
        FROM FeePayments
        GROUP BY student_fee_id
    ) fp ON sf.student_fee_id = fp.student_fee_id
    WHERE sf.fee_category_id = ?
    GROUP BY c.class_id, c.class_name_ar
    ORDER BY c.class_name_ar
    """, (category_id,))
    
    class_columns = [column[0] for column in cursor.description]
    class_collections = []
    for row in cursor.fetchall():
        item = dict(zip(class_columns, row))
        item['total_pending'] = item['total_amount'] - item['total_discounts'] - item['total_collected']
        item['collection_percentage'] = round((item['total_collected'] / item['total_amount'] * 100), 1) if item['total_amount'] > 0 else 0
        class_collections.append(item)
    
    # Get recent payments for this category
    cursor.execute("""
    SELECT TOP 20
        fp.payment_id,
        fp.payment_date,
        fp.amount_paid,
        fp.receipt_number,
        s.student_id,
        s.student_number,
        CONCAT(s.first_name_ar, ' ', s.last_name_ar) as student_name,
        c.class_name_ar
    FROM FeePayments fp
    JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
    JOIN Students s ON sf.student_id = s.student_id
    JOIN Classes c ON s.current_class_id = c.class_id
    WHERE sf.fee_category_id = ?
    ORDER BY fp.payment_date DESC, fp.payment_id DESC
    """, (category_id,))
    
    payment_columns = [column[0] for column in cursor.description]
    recent_payments = []
    for row in cursor.fetchall():
        recent_payments.append(dict(zip(payment_columns, row)))
    
    cursor.close()
    
    return render_template('finance/fee_category_collections.html',
                         category=category_dict,
                         stats=stats,
                         class_collections=class_collections,
                         recent_payments=recent_payments)
@finance_bp.route('/receipts/archive')
@role_required('admin', 'accountant')
def receipt_archive():
    """View archive of all receipts"""
    
    from datetime import datetime, timedelta
    
    # Get filter parameters
    page = request.args.get('page', 1, type=int)
    per_page = 20
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')
    search = request.args.get('search', '')
    
    # Set default dates if not provided
    if not from_date:
        from_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not to_date:
        to_date = datetime.now().strftime('%Y-%m-%d')
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        # Calculate TOTAL STATISTICS for all receipts (unfiltered)
        cursor.execute("""
        SELECT 
            COUNT(*) as total_receipts,
            ISNULL(SUM(amount_paid), 0) as total_amount,
            MAX(payment_date) as latest_date
        FROM FeePayments
        """)
        stats_row = cursor.fetchone()
        
        total_receipts_all = stats_row[0] or 0
        total_amount_all = float(stats_row[1] or 0)
        latest_date_all = stats_row[2]
        if latest_date_all:
            if hasattr(latest_date_all, 'strftime'):
                latest_date_all = latest_date_all.strftime('%Y-%m-%d')
            else:
                latest_date_all = str(latest_date_all)
        else:
            latest_date_all = '---'
        
        # Build the WHERE clause for filtered results
        where_clause = "WHERE 1=1"
        params = []
        
        # Add date filter
        where_clause += " AND fp.payment_date BETWEEN ? AND ?"
        params.extend([from_date, to_date])
        
        # Add search filter
        if search:
            where_clause += " AND (fp.receipt_number LIKE ? OR s.first_name_ar LIKE ? OR s.last_name_ar LIKE ? OR s.student_number LIKE ?)"
            search_param = f'%{search}%'
            params.extend([search_param, search_param, search_param, search_param])
        
        # Get total count for pagination (filtered)
        count_query = f"""
        SELECT COUNT(*) 
        FROM FeePayments fp
        JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
        JOIN Students s ON sf.student_id = s.student_id
        {where_clause}
        """
        cursor.execute(count_query, params)
        total_filtered = cursor.fetchone()[0]
        
        # Get receipts for current page (filtered)
        query = f"""
        SELECT 
            fp.payment_id,
            CONVERT(varchar, fp.payment_date, 23) as payment_date,
            fp.amount_paid as amount,
            fp.receipt_number,
            fp.payment_method,
            ISNULL(fp.notes, '') as notes,
            s.student_id,
            s.student_number,
            CONCAT(s.first_name_ar, ' ', s.last_name_ar) as student_name,
            c.class_name_ar,
            fc.category_name_ar as fee_category,
            u.full_name_ar as received_by_name
        FROM FeePayments fp
        JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
        JOIN Students s ON sf.student_id = s.student_id
        JOIN Classes c ON s.current_class_id = c.class_id
        JOIN FeeCategories fc ON sf.fee_category_id = fc.fee_category_id
        JOIN Users u ON fp.received_by = u.user_id
        {where_clause}
        ORDER BY fp.payment_date DESC, fp.payment_id DESC
        OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """
        
        pagination_params = params + [(page - 1) * per_page, per_page]
        cursor.execute(query, pagination_params)
        
        receipts = []
        if cursor.description:
            columns = [column[0] for column in cursor.description]
            for row in cursor.fetchall():
                receipt = {}
                for i, col in enumerate(columns):
                    receipt[col] = row[i]
                receipts.append(receipt)
        
        # Calculate filtered total amount
        total_amount_query = f"""
        SELECT ISNULL(SUM(fp.amount_paid), 0)
        FROM FeePayments fp
        JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
        JOIN Students s ON sf.student_id = s.student_id
        {where_clause}
        """
        cursor.execute(total_amount_query, params)
        total_amount_filtered = float(cursor.fetchone()[0] or 0)
        
        pagination = {
            'page': page,
            'per_page': per_page,
            'total': total_filtered,
            'pages': (total_filtered + per_page - 1) // per_page if total_filtered > 0 else 1
        }
        
        return render_template('finance/receipt_archive.html',
                             receipts=receipts,
                             pagination=pagination,
                             total_receipts=total_receipts_all,
                             total_amount=total_amount_all,
                             latest_date=latest_date_all,
                             filtered_total=total_filtered,
                             filtered_amount=total_amount_filtered,
                             from_date=from_date,
                             to_date=to_date,
                             search=search)
    
    except Exception as e:
        print(f"Error in receipt_archive: {e}")
        import traceback
        traceback.print_exc()
        flash('Error loading receipts archive', 'danger')
        return redirect(url_for('finance.dashboard'))
    
    finally:
        cursor.close()
@finance_bp.route('/receipt/<int:payment_id>/details')
def receipt_details(payment_id):
    """Get receipt details as JSON for modal"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        # Get receipt details - FIXED: Properly format the date
        cursor.execute("""
        SELECT 
            fp.payment_id,
            fp.payment_date,
            fp.amount_paid,
            fp.receipt_number,
            fp.payment_method,
            ISNULL(fp.notes, '') as notes,
            s.student_id,
            s.student_number,
            CONCAT(s.first_name_ar, ' ', s.last_name_ar) as student_name,
            ISNULL(c.class_name_ar, '') as class_name_ar,
            ISNULL(fc.category_name_ar, '') as fee_category,
            ISNULL(u.full_name_ar, '') as received_by,
            ISNULL(ss.setting_value, N'المدرسة') as school_name
        FROM FeePayments fp
        JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
        JOIN Students s ON sf.student_id = s.student_id
        LEFT JOIN Classes c ON s.current_class_id = c.class_id
        JOIN FeeCategories fc ON sf.fee_category_id = fc.fee_category_id
        JOIN Users u ON fp.received_by = u.user_id
        LEFT JOIN SystemSettings ss ON ss.setting_key = 'school_name_ar'
        WHERE fp.payment_id = ?
        """, (payment_id,))
        
        row = cursor.fetchone()
        cursor.close()
        
        if not row:
            return jsonify({'error': 'Receipt not found'}), 404
        
        # Format the date properly
        payment_date = row[1]
        if hasattr(payment_date, 'strftime'):
            # It's a date/datetime object
            formatted_date = payment_date.strftime('%Y-%m-%d')
        else:
            # It's already a string
            formatted_date = str(payment_date)
        
        # Build response with formatted date
        return jsonify({
            'payment_id': row[0],
            'payment_date': formatted_date,
            'amount': f"{float(row[2]):,.2f}",
            'receipt_number': row[3],
            'payment_method': row[4],
            'notes': row[5],
            'student_id': row[6],
            'student_number': row[7],
            'student_name': row[8],
            'class_name': row[9],
            'fee_category': row[10],
            'received_by': row[11],
            'school_name': row[12]
        })
        
    except Exception as e:
        print(f"Error in receipt_details: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
@finance_bp.route('/receipts/export')
@role_required('admin', 'accountant')
def export_receipts():
    """Export receipts list to CSV"""
    
    from datetime import datetime
    import csv
    import io
    
    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')
    search = request.args.get('search', '')
    
    db = get_db()
    cursor = db.cursor()
    
    query = """
    SELECT 
        fp.receipt_number,
        fp.payment_date,
        s.student_number,
        CONCAT(s.first_name_ar, ' ', s.last_name_ar) as student_name,
        fc.category_name_ar as fee_category,
        fp.amount_paid,
        fp.payment_method,
        u.full_name_ar as received_by
    FROM FeePayments fp
    JOIN StudentFees sf ON fp.student_fee_id = sf.student_fee_id
    JOIN Students s ON sf.student_id = s.student_id
    JOIN FeeCategories fc ON sf.fee_category_id = fc.fee_category_id
    JOIN Users u ON fp.received_by = u.user_id
    WHERE 1=1
    """
    params = []
    
    if from_date and to_date:
        query += " AND fp.payment_date BETWEEN ? AND ?"
        params.extend([from_date, to_date])
    
    if search:
        query += " AND (fp.receipt_number LIKE ? OR s.first_name_ar LIKE ? OR s.last_name_ar LIKE ?)"
        search_param = f'%{search}%'
        params.extend([search_param, search_param, search_param])
    
    query += " ORDER BY fp.payment_date DESC"
    
    cursor.execute(query, params)
    receipts = cursor.fetchall()
    cursor.close()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write headers
    writer.writerow(['رقم الإيصال', 'التاريخ', 'الرقم الدراسي', 'اسم الطالب', 'نوع الرسوم', 'المبلغ', 'طريقة الدفع', 'المستلم'])
    
    # Write data
    for r in receipts:
        writer.writerow(r)
    
    output.seek(0)
    
    filename = f"receipts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )