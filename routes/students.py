import csv
import io
import pandas as pd
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify,send_file
from models.student import Student
from models.class_ import Class
from models.audit import AuditLog
from database.db_config import get_db
from utils.file_upload import save_uploaded_file, delete_file
from datetime import datetime
from flask import send_file
from utils.permission_decorator import role_required

# Define the blueprint
students_bp = Blueprint('students', __name__)

@students_bp.route('/')
def list_students():
    """List all students"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    db = get_db()
    cursor = db.cursor()
    
    # Get all students with class information
    cursor.execute("""
    SELECT s.*, c.class_name_ar, c.class_name_en, c.class_id
    FROM Students s
    LEFT JOIN Classes c ON s.current_class_id = c.class_id
    WHERE s.status = 'active'
    ORDER BY s.student_id DESC
    """)
    
    columns = [column[0] for column in cursor.description]
    students = []
    
    for row in cursor.fetchall():
        students.append(dict(zip(columns, row)))
    
    # Calculate statistics
    total_students = len(students)
    active_students = sum(1 for s in students if s.get('status') == 'active')
    male_students = sum(1 for s in students if s.get('gender') == 'male')
    female_students = sum(1 for s in students if s.get('gender') == 'female')
    
    # Get unique classes for filter
    cursor.execute("""
    SELECT DISTINCT c.class_id, c.class_name_ar 
    FROM Classes c
    JOIN Students s ON c.class_id = s.current_class_id
    WHERE s.status = 'active'
    ORDER BY c.class_name_ar
    """)
    
    classes = []
    for row in cursor.fetchall():
        classes.append({'class_id': row[0], 'class_name_ar': row[1]})
    
    cursor.close()
    
    # Use the modern template
    return render_template('students/list.html',
                         students=students,
                         total_students=total_students,
                         active_students=active_students,
                         male_students=male_students,
                         female_students=female_students,
                         classes=classes)

@students_bp.route('/register', methods=['GET', 'POST'])
@role_required('admin')
def register():
    """Register new student"""
    
    if request.method == 'POST':
        try:
            # Get current academic year
            db = get_db()
            cursor = db.cursor()
            cursor.execute("SELECT year_id FROM AcademicYears WHERE is_current = 1")
            academic_year = cursor.fetchone()
            
            if not academic_year:
                flash('No active academic year found', 'danger')
                return redirect(url_for('students.register'))
            
            # Prepare student data
            student_data = {
                'first_name_ar': request.form['first_name_ar'],
                'last_name_ar': request.form['last_name_ar'],
                'first_name_en': request.form.get('first_name_en', ''),
                'last_name_en': request.form.get('last_name_en', ''),
                'birth_date': request.form['birth_date'],
                'gender': request.form['gender'],
                'nationality': request.form.get('nationality', 'Sudanese'),
                'address': request.form['address'],
                'phone': request.form['phone'],
                'email': request.form.get('email', ''),
                'enrollment_date': request.form['enrollment_date'],
                'class_id': request.form['class_id'],
                'academic_year_id': academic_year[0],
                'guardian_name_ar': request.form.get('guardian_name_ar'),
                'guardian_name_en': request.form.get('guardian_name_en'),
                'guardian_relation_ar': request.form.get('guardian_relation_ar'),
                'guardian_relation_en': request.form.get('guardian_relation_en'),
                'guardian_phone': request.form.get('guardian_phone'),
                'guardian_occupation': request.form.get('guardian_occupation'),
                'guardian_address': request.form.get('guardian_address')
            }
            
            # Create student
            student_id = Student.create(student_data)
            flash('Student registered successfully!', 'success')
            return redirect(url_for('students.view_student_profile', student_id=student_id))
            
        except Exception as e:
            flash(f'Error registering student: {str(e)}', 'danger')
            print(f"Error: {e}")
    
    # Get classes for dropdown
    classes = Class.get_all_active()
    
    # Get current date for default values
    today = datetime.now().strftime('%Y-%m-%d')
    
    return render_template('students/register.html', classes=classes, today=today)

@students_bp.route('/<int:student_id>')
def view_student(student_id):
    """View student details (legacy)"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    student = Student.get_by_id(student_id)
    if not student:
        flash('Student not found', 'danger')
        return redirect(url_for('students.list_students'))
    
    guardians = Student.get_guardians(student_id)
    return render_template('students/profile.html', student=student, guardians=guardians)

@students_bp.route('/<int:student_id>/profile')
def view_student_profile(student_id):
    """View enhanced student profile"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    student = Student.get_by_id(student_id)
    if not student:
        flash('Student not found', 'danger')
        return redirect(url_for('students.list_students'))
    
    guardians = Student.get_guardians(student_id)
    fee_summary = Student.get_fee_summary(student_id)
    attendance_summary = Student.get_attendance_summary(student_id)
    
    return render_template('students/profile.html',
                         student=student,
                         guardians=guardians,
                         fee_summary=fee_summary,
                         attendance_summary=attendance_summary)

@students_bp.route('/<int:student_id>/edit', methods=['GET', 'POST'])
@role_required('admin')
def edit_student(student_id):
    """Edit student information"""
 
    student = Student.get_by_id(student_id)
    if not student:
        flash('Student not found', 'danger')
        return redirect(url_for('students.list_students'))
    
    if request.method == 'POST':
        try:
            student_data = {
                'first_name_ar': request.form['first_name_ar'],
                'last_name_ar': request.form['last_name_ar'],
                'first_name_en': request.form.get('first_name_en', ''),
                'last_name_en': request.form.get('last_name_en', ''),
                'birth_date': request.form['birth_date'],
                'gender': request.form['gender'],
                'nationality': request.form.get('nationality', 'Sudanese'),
                'address': request.form['address'],
                'phone': request.form['phone'],
                'email': request.form.get('email', ''),
                'class_id': request.form['class_id']
            }
            
            if Student.update(student_id, student_data):
                flash('Student updated successfully!', 'success')
                return redirect(url_for('students.view_student_profile', student_id=student_id))
            else:
                flash('No changes were made', 'info')
                
        except Exception as e:
            flash(f'Error updating student: {str(e)}', 'danger')
    
    classes = Class.get_all_active()
    return render_template('students/edit.html', student=student, classes=classes)
@students_bp.route('/<int:student_id>/update-contact', methods=['POST'])
@role_required('admin')
def update_contact(student_id):
    """Update student contact information"""
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("""
        UPDATE Students
        SET phone = ?, email = ?, address = ?
        WHERE student_id = ?
        """, (
            request.form.get('phone', ''),
            request.form.get('email', ''),
            request.form.get('address', ''),
            student_id
        ))
        db.commit()
        flash('Contact information updated successfully', 'success')
    except Exception as e:
        flash(f'Error updating contact: {str(e)}', 'danger')
    finally:
        cursor.close()
    
    return redirect(url_for('students.edit_student', student_id=student_id))

@students_bp.route('/<int:student_id>/update-academic', methods=['POST'])
@role_required('admin')
def update_academic(student_id):
    """Update student academic information"""
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute("""
        UPDATE Students
        SET current_class_id = ?, enrollment_date = ?, status = ?
        WHERE student_id = ?
        """, (
            request.form['class_id'],
            request.form.get('enrollment_date'),
            request.form.get('status', 'active'),
            student_id
        ))
        db.commit()
        flash('Academic information updated successfully', 'success')
    except Exception as e:
        flash(f'Error updating academic info: {str(e)}', 'danger')
    finally:
        cursor.close()
    
    return redirect(url_for('students.edit_student', student_id=student_id))
@students_bp.route('/<int:student_id>/delete', methods=['POST'])
@role_required('admin')
def delete_student(student_id):
    """Delete (withdraw) a student"""
    
    # Get student data before deletion
    student = Student.get_by_id(student_id)
    student_name = f"{student['first_name_ar']} {student['last_name_ar']}" if student else 'Unknown'
    student_number = student['student_number'] if student else 'Unknown'
    
    if Student.delete(student_id):
        # ========== ADD AUDIT LOG HERE ==========
        from models.audit import AuditLog
        AuditLog.log_action(
            action_type='DELETE',
            table_name='Students',
            record_id=student_id,
            old_data={
                'name': student_name,
                'student_number': student_number,
                'class': student.get('class_name_ar') if student else 'Unknown'
            },
            description=f'Student withdrawn: {student_name} ({student_number})'
        )
        # ========== END AUDIT LOG ==========
        
        flash('Student withdrawn successfully', 'success')
    else:
        flash('Error withdrawing student', 'danger')
    
    return redirect(url_for('students.list_students'))

@students_bp.route('/<int:student_id>/generate-id')
def generate_id(student_id):
    """Generate student ID card"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    student = Student.get_by_id(student_id)
    if not student:
        flash('Student not found', 'danger')
        return redirect(url_for('students.list_students'))
    
    # For now, just flash a message (you can implement PDF generation later)
    flash('ID card generation feature coming soon!', 'info')
    return redirect(url_for('students.view_student_profile', student_id=student_id))

@students_bp.route('/<int:student_id>/upload-photo', methods=['POST'])
@role_required('admin')
def upload_photo(student_id):
    """Upload student profile photo"""
    
    if 'photo' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    
    file = request.files['photo']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    # Save the uploaded file
    file_path = save_uploaded_file(file, 'students')
    if not file_path:
        return jsonify({'success': False, 'error': 'Invalid file type'}), 400
    
    # Update database
    old_image = Student.update_image(student_id, file_path)
    
    # Delete old image if it exists
    if old_image:
        delete_file(old_image)
    
    return jsonify({'success': True, 'path': file_path})

@students_bp.route('/<int:student_id>/add-guardian', methods=['POST'])
@role_required('admin')
def add_guardian(student_id):
    """Add guardian to student"""

    
    try:
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("""
        INSERT INTO Guardians (
            student_id, full_name_ar, full_name_en, relationship_ar, relationship_en,
            phone, phone2, occupation, email, is_primary
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            student_id,
            request.form['full_name_ar'],
            request.form.get('full_name_en', ''),
            request.form['relationship_ar'],
            request.form.get('relationship_en', ''),
            request.form['phone'],
            request.form.get('phone2', ''),
            request.form.get('occupation', ''),
            request.form.get('email', ''),
            0  # Not primary by default
        ))
        
        db.commit()
        cursor.close()
        flash('Guardian added successfully', 'success')
        
    except Exception as e:
        flash(f'Error adding guardian: {str(e)}', 'danger')
    
    return redirect(url_for('students.view_student_profile', student_id=student_id))@students_bp.route('/export')
@students_bp.route('/export/csv')
def export_students_csv():
    """Export students to CSV"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    db = get_db()
    cursor = db.cursor()
    
    # Get filter parameters
    class_id = request.args.get('class_id')
    gender = request.args.get('gender')
    
    # Build query with filters
    query = """
    SELECT 
        s.student_number,
        s.first_name_ar,
        s.last_name_ar,
        s.first_name_en,
        s.last_name_en,
        ISNULL(c.class_name_ar, '') as class_name_ar,
        CASE WHEN s.gender = 'male' THEN 'ذكر' ELSE 'أنثى' END as gender,
        CONVERT(varchar, s.birth_date, 23) as birth_date,
        ISNULL(s.phone, '') as phone,
        ISNULL(s.email, '') as email,
        ISNULL(s.address, '') as address,
        CONVERT(varchar, s.enrollment_date, 23) as enrollment_date,
        CASE WHEN s.status = 'active' THEN 'نشط' ELSE 'غير نشط' END as status
    FROM Students s
    LEFT JOIN Classes c ON s.current_class_id = c.class_id
    WHERE s.status = 'active'
    """
    
    params = []
    if class_id:
        query += " AND s.current_class_id = ?"
        params.append(class_id)
    if gender:
        query += " AND s.gender = ?"
        params.append(gender)
    
    query += " ORDER BY s.first_name_ar"
    
    cursor.execute(query, params)
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write headers (Arabic)
    headers = [
        'الرقم الدراسي', 
        'الاسم الأول (عربي)', 
        'الاسم الأخير (عربي)',
        'الاسم الأول (إنجليزي)', 
        'الاسم الأخير (إنجليزي)', 
        'الصف',
        'الجنس', 
        'تاريخ الميلاد', 
        'رقم الهاتف', 
        'البريد الإلكتروني',
        'العنوان', 
        'تاريخ التسجيل', 
        'الحالة'
    ]
    writer.writerow(headers)
    
    # Write data
    row_count = 0
    for row in cursor.fetchall():
        writer.writerow(row)
        row_count += 1
    
    cursor.close()
    
    output.seek(0)
    
    filename = f"students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )
@students_bp.route('/export/excel')
def export_students_excel():
    """Export students to Excel"""
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    db = get_db()
    cursor = db.cursor()
    
    # Get filter parameters
    class_id = request.args.get('class_id')
    gender = request.args.get('gender')
    
    # Build query with filters
    query = """
    SELECT 
        s.student_number,
        s.first_name_ar,
        s.last_name_ar,
        s.first_name_en,
        s.last_name_en,
        ISNULL(c.class_name_ar, '') as class_name_ar,
        CASE WHEN s.gender = 'male' THEN 'ذكر' ELSE 'أنثى' END as gender,
        CONVERT(varchar, s.birth_date, 23) as birth_date,
        ISNULL(s.phone, '') as phone,
        ISNULL(s.email, '') as email,
        ISNULL(s.address, '') as address,
        CONVERT(varchar, s.enrollment_date, 23) as enrollment_date,
        CASE WHEN s.status = 'active' THEN 'نشط' ELSE 'غير نشط' END as status
    FROM Students s
    LEFT JOIN Classes c ON s.current_class_id = c.class_id
    WHERE s.status = 'active'
    """
    
    params = []
    if class_id:
        query += " AND s.current_class_id = ?"
        params.append(class_id)
    if gender:
        query += " AND s.gender = ?"
        params.append(gender)
    
    query += " ORDER BY s.first_name_ar"
    
    print(f"Executing query: {query}")
    print(f"Params: {params}")
    
    cursor.execute(query, params)
    
    # Fetch all data
    rows = cursor.fetchall()
    print(f"Fetched {len(rows)} rows")
    
    if not rows:
        print("No data found")
        flash('No students found to export', 'warning')
        return redirect(url_for('students.list_students'))
    
    # Get column names from cursor description
    columns = [column[0] for column in cursor.description]
    print(f"Columns: {columns}")
    
    # Convert rows to list of dictionaries
    data = []
    for row in rows:
        row_dict = {}
        for i, col in enumerate(columns):
            row_dict[col] = row[i]
        data.append(row_dict)
    
    # Create DataFrame
    df = pd.DataFrame(data)
    print(f"DataFrame shape: {df.shape}")
    
    # Rename columns to Arabic for better readability
    column_mapping = {
        'student_number': 'الرقم الدراسي',
        'first_name_ar': 'الاسم الأول (عربي)',
        'last_name_ar': 'الاسم الأخير (عربي)',
        'first_name_en': 'الاسم الأول (إنجليزي)',
        'last_name_en': 'الاسم الأخير (إنجليزي)',
        'class_name_ar': 'الصف',
        'gender': 'الجنس',
        'birth_date': 'تاريخ الميلاد',
        'phone': 'رقم الهاتف',
        'email': 'البريد الإلكتروني',
        'address': 'العنوان',
        'enrollment_date': 'تاريخ التسجيل',
        'status': 'الحالة'
    }
    
    df = df.rename(columns=column_mapping)
    
    # Create Excel file in memory
    output = io.BytesIO()
    
    # Use ExcelWriter to create formatted Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Students', index=False, startrow=2)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Students']
        
        # Add title
        worksheet.cell(row=1, column=1, value=f"تقرير الطلاب - {datetime.now().strftime('%Y-%m-%d')}")
        
        # Merge title cells
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        
        # Format title
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.font = Font(size=16, bold=True, color="875A7B")
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add generation info
        worksheet.cell(row=2, column=1, value=f"تم التصدير في: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))
        info_cell = worksheet.cell(row=2, column=1)
        info_cell.font = Font(size=10, italic=True, color="666666")
        info_cell.alignment = Alignment(horizontal='center')
        
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="875A7B", end_color="875A7B", fill_type="solid")
        
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=3, column=col_num)
            cell.value = column
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for col_num, column in enumerate(df.columns, 1):
            max_length = len(column)  # Start with header length
            column_letter = get_column_letter(col_num)
            
            # Check data in this column
            for row_num in range(4, 4 + len(df)):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    cursor.close()
    output.seek(0)
    
    filename = f"students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )